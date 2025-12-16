from fastapi import APIRouter, HTTPException, Path, Depends, Query
from fastapi.responses import StreamingResponse, PlainTextResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime
from bson import ObjectId
from db import db
from utils import serialize_doc, extract_final_xml
from models import UserInDB
from dateutil import parser
from routers.auth import get_current_user
import xml.etree.ElementTree as ET
import io
import re
import json
import base64
import zipfile

router = APIRouter(prefix="/tasks", tags=["Tasks"])


@router.get("")
def list_tasks(
        emitterId: str | None = None,
        status: str | None = None,
        mes: int | None = Query(None, ge=1, le=12),
        ano: int | None = Query(None, ge=2000),
        current_user: UserInDB = Depends(get_current_user)
):
    user_id = ObjectId(current_user.id)
    q = {"user_id": user_id}

    if emitterId:
        q["emitter_id"] = emitterId
    if status:
        q["status"] = status

    if mes and ano:
        inicio = datetime(ano, mes, 1)
        fim = datetime(ano, mes + 1, 1) if mes < 12 else datetime(ano + 1, 1, 1)

        q["competencia"] = {
            "$gte": inicio.strftime("%Y-%m-%d"),
            "$lt": fim.strftime("%Y-%m-%d"),
        }

    pipeline = [
        {"$match": q},
        {"$sort": {"created_at": -1}},

        # --- Tratamento de IDs ---
        {"$addFields": {
            "client_id_obj": {
                "$cond": {
                    "if": {
                        "$and": [
                            {"$ne": ["$client_id", None]},
                            {"$regexMatch": {"input": "$client_id", "regex": "^[0-9a-fA-F]{24}$"}}
                        ]
                    },
                    "then": {"$toObjectId": "$client_id"},
                    "else": None
                }
            },
            "emitter_id_obj": {
                "$cond": {
                    "if": {
                        "$and": [
                            {"$ne": ["$emitter_id", None]},
                            {"$regexMatch": {"input": "$emitter_id", "regex": "^[0-9a-fA-F]{24}$"}}
                        ]
                    },
                    "then": {"$toObjectId": "$emitter_id"},
                    "else": None
                }
            }
        }},

        # --- Lookups ---
        {"$lookup": {
            "from": "clients",
            "localField": "client_id_obj",
            "foreignField": "_id",
            "as": "cliente"
        }},
        {"$unwind": {"path": "$cliente", "preserveNullAndEmptyArrays": True}},

        {"$lookup": {
            "from": "emitters",
            "localField": "emitter_id_obj",
            "foreignField": "_id",
            "as": "emissor"
        }},
        {"$unwind": {"path": "$emissor", "preserveNullAndEmptyArrays": True}},

        # --- OTIMIZAÇÃO DE PERFORMANCE (NOVO) ---
        # 1. Calcula se tem PDF antes de remover o campo pesado
        {"$addFields": {
            "has_pdf_temp": {
                "$cond": {
                    "if": {
                        "$and": [
                            {"$ifNull": ["$transmit.pdf_base64", False]},
                            {"$ne": ["$transmit.pdf_base64", ""]}
                        ]
                    },
                    "then": True,
                    "else": False
                }
            }
        }},

        # 2. Remove os campos gigantes do retorno da query
        {"$project": {
            "transmit.pdf_base64": 0,
            "transmit.raw_response": 0,
            "transmit.receipt.bruto": 0,
            "transmit.receipt.xml_nfse": 0,
            "transmit.xml_nfse": 0,
            "response.xml": 0,
            # Mantemos o resto
        }}
    ]

    cur = db.tasks.aggregate(pipeline)
    out = []

    for t in cur:
        t = serialize_doc(t)

        # --- Fallback seguro: busca cliente direto se lookup falhou
        if not t.get("cliente") and t.get("client_id"):
            try:
                if ObjectId.is_valid(t["client_id"]):
                    cliente = db.clients.find_one({"_id": ObjectId(t["client_id"]), "user_id": user_id})
                else:
                    cliente = db.clients.find_one({"_id": t["client_id"], "user_id": user_id})
            except Exception:
                cliente = None
            if cliente:
                t["cliente"] = serialize_doc(cliente)

        # --- Preenche campos de exibição
        t["cliente_nome"] = t.get("cliente", {}).get("nome") or "-"
        t["cliente_email"] = t.get("cliente", {}).get("email") or "-"
        t["emissor_nome"] = t.get("emissor", {}).get("razaoSocial") or "-"

        # --- OTIMIZAÇÃO: Usa o booleano calculado no Mongo, pois o base64 não existe mais aqui
        t["has_pdf"] = t.get("has_pdf_temp", False)
        t.pop("has_pdf_temp", None)
        t.pop("cliente", None)
        t.pop("emissor", None)
        out.append(t)

    return out


@router.get("/resumo")
def resumo_por_emissor(
        mes: int = Query(..., ge=1, le=12),
        ano: int = Query(..., ge=2000),
        current_user: UserInDB = Depends(get_current_user)
):
    """Resumo de notas por emissor, filtrado por mês/ano e organização."""
    user_id = ObjectId(current_user.id)
    inicio = datetime(ano, mes, 1)
    fim = datetime(ano, mes + 1, 1) if mes < 12 else datetime(ano + 1, 1, 1)

    pipeline = [
        # FILTRO DE SEGURANÇA: Garante que o resumo é apenas da organização do usuário
        {"$match": {
            "user_id": user_id,
            "status": {"$in": ["accepted", "pending"]},
            "$or": [
                {"competencia": {"$gte": inicio.strftime("%Y-%m-%d"), "$lt": fim.strftime("%Y-%m-%d")}},
                {"competencia": {"$gte": inicio, "$lt": fim}},
            ]
        }},
        {"$addFields": {
            "emitter_id_obj": {"$toObjectId": "$emitter_id"}
        }},
        {"$lookup": {
            "from": "emitters", "localField": "emitter_id_obj",
            "foreignField": "_id", "as": "emissor"
        }},
        {"$unwind": {"path": "$emissor", "preserveNullAndEmptyArrays": True}},
        {"$group": {
            "_id": "$emissor._id",
            "emissor_nome": {"$first": "$emissor.razaoSocial"},
            "total_notas": {"$sum": 1},
            "valor_total": {"$sum": {"$ifNull": ["$response.valor", "$valor"]}},
        }},
        {"$sort": {"valor_total": -1}}
    ]

    return [serialize_doc(r) for r in db.tasks.aggregate(pipeline)]


# ------------------------------------------------
@router.get("/batch/xml")
def download_all_xml(emitterId: str | None = None, current_user: UserInDB = Depends(get_current_user)):
    user_id = ObjectId(current_user.id)
    q = {"user_id": user_id, "status": "accepted"}
    if emitterId:
        q["emitter_id"] = emitterId

    cur = db.tasks.find(q)
    mem = io.BytesIO()

    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as zf:
        for t in cur:
            tr = t.get("transmit") or {}
            xml_final = _pick_final_xml_from_transmit(tr, t)
            if not xml_final:
                continue
            zf.writestr(f"nfse_{t['_id']}.xml", xml_final)

    mem.seek(0)
    return StreamingResponse(
        mem,
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="xml.zip"'}
    )


@router.get("/batch/pdf")
def download_all_pdf(emitterId: str | None = None, current_user: UserInDB = Depends(get_current_user)):
    """
    Baixa todas as DANFSe (PDFs) das notas aceitas.
    """
    user_id = ObjectId(current_user.id)
    q = {"user_id": user_id, "status": "accepted"}

    if emitterId:
        q["emitter_id"] = emitterId

    cur = db.tasks.find(q)
    mem = io.BytesIO()

    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as zf:
        for t in cur:
            tr = t.get("transmit") or {}
            pdf_b64 = tr.get("pdf_base64")
            if not pdf_b64:
                continue  # pula notas sem PDF ainda

            try:
                pdf_bytes = base64.b64decode(pdf_b64)
                filename = f"danfs_{t['_id']}.pdf"
                zf.writestr(filename, pdf_bytes)
            except Exception:
                continue

    mem.seek(0)
    return StreamingResponse(
        mem,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=danfs.zip"}
    )


# -------------------------------------------------
@router.get("/{task_id}/xml")
def download_xml(task_id: str = Path(...), current_user: UserInDB = Depends(get_current_user)):
    """
    Baixa o XML FINAL autorizado da NFSe (proveniente de transmit.*),
    com múltiplos fallbacks (xml_nfse, nfseXmlGZipB64, raw XML, extract_final_xml).
    """
    user_id = ObjectId(current_user.id)
    task = db.tasks.find_one({"_id": ObjectId(task_id), "user_id": user_id})
    if not task:
        raise HTTPException(status_code=404, detail="Task não encontrada")

    tr = task.get("transmit") or {}
    xml_nfse = _pick_final_xml_from_transmit(tr)
    if not xml_nfse:
        raise HTTPException(status_code=400, detail="Task ainda não possui XML final para download.")

    return PlainTextResponse(
        content=xml_nfse,
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="nfse_{task_id}.xml"'}
    )


@router.get("/{task_id}/guia")
def download_guia(task_id: str = Path(...), current_user: UserInDB = Depends(get_current_user)):
    """
    Baixa a guia oficial em PDF **se** a prefeitura forneceu (transmit.pdf_base64).
    Caso não exista, retorna 400 (sem gerar PDF fake).
    """
    user_id = ObjectId(current_user.id)
    task = db.tasks.find_one({"_id": ObjectId(task_id), "user_id": user_id})
    if not task:
        raise HTTPException(status_code=404, detail="Task não encontrada")

    tr = task.get("transmit") or {}
    pdf_b64 = tr.get("pdf_base64")

    if not pdf_b64:
        raise HTTPException(status_code=400, detail="Guia oficial ainda não disponível para esta nota.")

    try:
        pdf_bytes = base64.b64decode(pdf_b64)
    except Exception:
        raise HTTPException(status_code=400, detail="PDF inválido no retorno da prefeitura.")

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="danfs_{task_id}.pdf"'}
    )


@router.delete("/{task_id}")
def delete_task(task_id: str, current_user: UserInDB = Depends(get_current_user)):
    """
    Exclui uma task (geralmente usada para limpar erros resolvidos).
    """
    user_id = ObjectId(current_user.id)
    query = {"_id": ObjectId(task_id), "user_id": user_id}

    result = db.tasks.delete_one(query)

    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Task não encontrada ou já removida")

    return {"msg": "Task descartada com sucesso"}


def _pick_final_xml_from_transmit(tr: dict, task: dict | None = None) -> str | None:
    """
    Retorna o XML final da NFSe, com todos os fallbacks possíveis.
    """
    if not tr:
        return None

    # 1) XML final puro já pronto
    xml_nfse = tr.get("xml_nfse")
    if xml_nfse:
        return xml_nfse

    # ✅ 2) Fallback legado: response.xml (unitário usa isso)
    xml_legacy = (task.get("response") or {}).get("xml") if task else None
    if xml_legacy:
        return xml_legacy

    raw = tr.get("raw_response")
    if not raw:
        return None

    # 3) JSON com nfseXmlGZipB64 → descompacta
    try:
        data = json.loads(raw)
        gz_b64 = data.get("nfseXmlGZipB64") or data.get("nfseXmlGzipB64")
        if gz_b64:
            from utils import nfse_gzip_b64_to_xml
            xml = nfse_gzip_b64_to_xml(gz_b64)
            if xml:
                return xml
    except Exception:
        pass

    # 4) Se já veio XML puro no raw_response
    if isinstance(raw, str) and raw.strip().startswith("<"):
        return raw

    # 5) fallback final para casos especiais
    return extract_final_xml(raw)


@router.get("/export")
def export_xlsx(
        mes: int = Query(..., ge=1, le=12),
        ano: int = Query(..., ge=2000),
        emitterId: str | None = None,
        current_user: UserInDB = Depends(get_current_user)
):
    user_id = ObjectId(current_user.id)

    # --- Filtros ---
    inicio = datetime(ano, mes, 1)
    fim = datetime(ano + (1 if mes == 12 else 0), (mes % 12) + 1, 1)

    filtro = {
        "user_id": user_id,
        "$or": [
            {"competencia": {"$gte": inicio.strftime("%Y-%m-%d"),
                             "$lt": fim.strftime("%Y-%m-%d")}},
            {"competencia": {"$gte": inicio, "$lt": fim}},
        ]
    }

    if emitterId:
        filtro["emitter_id"] = emitterId

    # --- Pipeline ---
    pipeline = [
        {"$match": filtro},
        {"$addFields": {
            "client_id_obj": {
                "$cond": {
                    "if": {"$regexMatch": {"input": "$client_id", "regex": "^[0-9a-fA-F]{24}$"}},
                    "then": {"$toObjectId": "$client_id"},
                    "else": None
                }
            },
            "emitter_id_obj": {
                "$cond": {
                    "if": {"$regexMatch": {"input": "$emitter_id", "regex": "^[0-9a-fA-F]{24}$"}},
                    "then": {"$toObjectId": "$emitter_id"},
                    "else": None
                }
            },
            "draft_id_obj": {
                "$cond": {
                    "if": {
                        "$and": [
                            {"$ne": ["$source.draft_id", None]},
                            {"$regexMatch": {"input": "$source.draft_id", "regex": "^[0-9a-fA-F]{24}$"}}
                        ]
                    },
                    "then": {"$toObjectId": "$source.draft_id"},
                    "else": None
                }
            }
        }},
        {"$lookup": {"from": "clients", "localField": "client_id_obj", "foreignField": "_id", "as": "cliente"}},
        {"$unwind": {"path": "$cliente", "preserveNullAndEmptyArrays": True}},
        {"$lookup": {"from": "emitters", "localField": "emitter_id_obj", "foreignField": "_id", "as": "emissor"}},
        {"$unwind": {"path": "$emissor", "preserveNullAndEmptyArrays": True}},
        {"$lookup": {"from": "tasks_draft", "localField": "draft_id_obj", "foreignField": "_id", "as": "draft"}},
        {"$unwind": {"path": "$draft", "preserveNullAndEmptyArrays": True}},
        {"$sort": {"emissor.razaoSocial": 1, "created_at": 1}}
    ]

    cur = db.tasks.aggregate(pipeline)

    # --- Configuração do Excel ---
    wb = openpyxl.Workbook()
    if wb.active:
        wb.remove(wb.active)

    HEADER = [
        "STATUS", "CHAVE DE ACESSO", "Nº DPS", "Nº NFSe", "DATA DE ENVIO",
        "DATA DE CANCELAMENTO", "VALOR", "ALÍQUOTA (%)", "PRESTADOR",
        "CNPJ PRESTADOR", "TOMADOR", "CNPJ/CPF TOMADOR", "REGIME DE TRIBUTAÇÃO",
        "NATUREZA DA OPERAÇÃO", "ISS RETIDO", "DESCRIÇÃO DOS SERVIÇOS",
        "CÓDIGO DA ATIVIDADE", "CIDADE TOMADOR", "ENDEREÇO TOMADOR"
    ]

    # Estilos
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_border = Border(
        left=Side(border_style="thin"), right=Side(border_style="thin"),
        top=Side(border_style="thin"), bottom=Side(border_style="thin"),
    )
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center")

    sheets_map = {}

    def get_or_create_sheet(emissor_nome):
        clean_name = re.sub(r'[\\/*?:\[\]]', '', emissor_nome or "Sem Emissor")[:30]
        if clean_name not in sheets_map:
            ws = wb.create_sheet(title=clean_name)
            ws.append(HEADER)
            for col in range(1, len(HEADER) + 1):
                c = ws.cell(row=1, column=col)
                c.font = header_font
                c.fill = header_fill
                c.border = header_border
                c.alignment = header_align
            sheets_map[clean_name] = ws
        return sheets_map[clean_name]

    # --- Helpers ---
    def format_activity_code(code):
        if not code:
            return ""
        clean = re.sub(r'\D', '', str(code))
        if len(clean) == 6:
            return f"{clean[:2]}.{clean[2:4]}.{clean[4:]}"
        return str(code)

    def safe_date(val):
        if not val: return None
        if isinstance(val, datetime):
            return val.replace(tzinfo=None)
        try:
            dt = parser.parse(str(val))
            return dt.replace(tzinfo=None)
        except:
            return str(val)

    def safe_float(val):
        if val is None or val == "":
            return None
        try:
            return float(val)
        except:
            return None

    # --- PARSER ROBUSTO (XML + REGEX) ---
    def parse_nfse_data(xml_content):
        """
        Extrai dados do XML usando ElementTree.
        Se falhar ou não encontrar campos chave, usa Regex como fallback.
        """
        data = {}
        if not xml_content:
            return data

        # 1. Tentativa Estruturada (ElementTree)
        try:
            ns = {"n": "http://www.sped.fazenda.gov.br/nfse"}
            # Remove possíveis caracteres BOM ou sujeira antes do <
            clean_xml = xml_content[xml_content.find("<"):]
            root = ET.fromstring(clean_xml)

            # Busca com namespace
            aliq_tag = root.find(".//n:pTotTribSN", ns) or root.find(".//n:pAliq", ns) or root.find(".//n:Aliquota", ns)
            if aliq_tag is not None:
                data['aliquota'] = aliq_tag.text

            vserv_tag = root.find(".//n:vServ", ns)
            if vserv_tag is not None:
                data['valor'] = vserv_tag.text

            iss_tag = root.find(".//n:tpRetISSQN", ns)
            if iss_tag is not None:
                data['iss_retido'] = (iss_tag.text == "1")

            desc_tag = root.find(".//n:xDescServ", ns)
            if desc_tag is not None:
                data['descricao'] = desc_tag.text

            cod_tag = root.find(".//n:cTribNac", ns)
            if cod_tag is not None:
                data['cod_servico'] = cod_tag.text

            nat_desc = root.find(".//n:xTribNac", ns)
            if nat_desc is not None:
                data['natureza'] = nat_desc.text

            toma = root.find(".//n:toma", ns)
            if toma is not None:
                xNome = toma.find(".//n:xNome", ns)
                if xNome is not None:
                    data['tomador_nome'] = xNome.text

                cnpj = toma.find(".//n:CNPJ", ns)
                cpf = toma.find(".//n:CPF", ns)
                data['tomador_doc'] = (cnpj.text if cnpj is not None else (cpf.text if cpf is not None else ""))

                end = toma.find(".//n:end", ns)
                if end is not None:
                    lgr = end.find(".//n:xLgr", ns)
                    nro = end.find(".//n:nro", ns)
                    bairro = end.find(".//n:xBairro", ns)
                    mun = end.find(".//n:xMun", ns)

                    lgr_t = lgr.text if lgr is not None else ""
                    nro_t = nro.text if nro is not None else ""
                    bairro_t = bairro.text if bairro is not None else ""
                    data['tomador_end'] = f"{lgr_t} {nro_t} {bairro_t}".strip()
                    if mun is not None: data['tomador_cidade'] = mun.text

        except Exception:
            # Se der erro no parse estruturado, não faz nada e deixa o regex salvar
            pass

        # 2. Fallback Bruto (Regex) - "Salva-vidas"
        # Se a alíquota não foi encontrada via estrutura, busca no texto bruto
        if data.get('aliquota') is None:
            match = re.search(r'<(?:pTotTribSN|pAliq|Aliquota)>([\d\.]+)</', xml_content)
            if match:
                data['aliquota'] = match.group(1)

        if data.get('descricao') is None:
            match = re.search(r'<xDescServ>(.*?)</xDescServ>', xml_content, re.DOTALL | re.IGNORECASE)
            if match:
                data['descricao'] = match.group(1)

        if data.get('valor') is None:
            match = re.search(r'<vServ>([\d\.]+)</', xml_content)
            if match:
                data['valor'] = match.group(1)

        if data.get('iss_retido') is None:
            if "<tpRetISSQN>1</tpRetISSQN>" in xml_content:
                data['iss_retido'] = True
            elif "<tpRetISSQN>2</tpRetISSQN>" in xml_content:
                data['iss_retido'] = False

        return data

    has_data = False

    for t in cur:
        has_data = True
        t = serialize_doc(t)

        tr = t.get("transmit") or {}
        receipt = tr.get("receipt") or {}
        dps_resumo = t.get("dps") or {}
        cliente = t.get("cliente") or {}
        emissor = t.get("emissor") or {}
        draft = t.get("draft") or {}

        ws = get_or_create_sheet(emissor.get("razaoSocial"))

        xml_text = (tr.get("xml_nfse") or (t.get("response") or {}).get("xml"))

        # Extração de Dados
        xml_data = parse_nfse_data(xml_text)

        # --- Consolidação ---

        # Alíquota
        aliquota_val = safe_float(xml_data.get('aliquota'))
        if aliquota_val is None:
            aliquota_val = safe_float(t.get("aliquota"))
        if aliquota_val is None:
            aliquota_val = safe_float(draft.get("aliquota"))

        # Valor
        valor_val = safe_float(xml_data.get('valor'))
        if valor_val is None:
            valor_val = safe_float(t.get("valor"))

        # ISS Retido
        iss_retido_val = xml_data.get('iss_retido')
        if iss_retido_val is None:
            iss_retido_val = t.get("iss_retido")
        if iss_retido_val is None:
            iss_retido_val = draft.get("iss_retido")
        iss_retido_str = "Sim" if iss_retido_val is True else ("Não" if iss_retido_val is False else "")

        # Descrição
        descricao = xml_data.get('descricao')
        if not descricao:
            descricao = t.get("descricao")
        if not descricao:
            descricao = draft.get("descricao") or ""

        # Limpeza de quebras de linha (solicitado)
        if descricao:
            descricao = " ".join(descricao.split())

        # Código Serviço
        cod_servico_raw = xml_data.get('cod_servico')
        if not cod_servico_raw:
            cod_servico_raw = t.get("cod_servico")
        if not cod_servico_raw:
            cod_servico_raw = draft.get("cod_servico") or ""
        cod_servico_final = format_activity_code(cod_servico_raw)

        # Dados do Tomador
        tomador_nome = xml_data.get('tomador_nome') or cliente.get("nome") or ""
        tomador_doc = xml_data.get('tomador_doc') or cliente.get("cnpj") or cliente.get("cpf") or ""
        tomador_end = xml_data.get('tomador_end')
        if not tomador_end:
            tomador_end = f"{cliente.get('logradouro', '')} {cliente.get('numero', '')} {cliente.get('bairro', '')}".strip()
        tomador_cidade = xml_data.get('tomador_cidade') or cliente.get("cidade") or ""

        natureza_texto = xml_data.get('natureza') or ""

        regime_raw = emissor.get("regimeTributacao", "")
        regime_map = {
            "Simples Nacional": "Simples Nacional", "MEI": "MEI",
            "Lucro Presumido": "Lucro Presumido", "Lucro Real": "Lucro Real",
        }
        regime = regime_map.get(regime_raw, regime_raw)

        status_raw = t.get("status", "")
        status_map_label = {
            "accepted": "AUTORIZADA", "rejected": "REJEITADA",
            "canceled": "CANCELADA", "error": "ERRO",
            "pending": "PENDENTE", "processing": "PROCESSANDO"
        }
        status_final = status_map_label.get(status_raw, status_raw.upper())

        data_envio = safe_date(t.get("sent_at") or t.get("created_at"))
        data_canc = safe_date(t.get("canceled_at"))

        row = [
            status_final,
            tr.get("chave_acesso") or receipt.get("chave_acesso") or "",
            dps_resumo.get("numero") or "",
            receipt.get("numero_nfse") or "",
            data_envio,
            data_canc,
            valor_val,
            aliquota_val,
            emissor.get("razaoSocial", ""),
            emissor.get("cnpj", ""),
            tomador_nome,
            tomador_doc,
            regime,
            natureza_texto,
            iss_retido_str,
            descricao,
            cod_servico_final,
            tomador_cidade,
            tomador_end
        ]

        ws.append(row)

        current_row = ws.max_row
        ws.cell(row=current_row, column=5).number_format = 'dd/mm/yyyy hh:mm:ss'
        ws.cell(row=current_row, column=6).number_format = 'dd/mm/yyyy hh:mm:ss'
        ws.cell(row=current_row, column=7).number_format = '#,##0.00'
        ws.cell(row=current_row, column=8).number_format = '0.00'
        # Removido alinhamento wrap_text

    if not has_data:
        wb.create_sheet("Sem Dados")

    for ws in wb.worksheets:
        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            length = 0
            for c in col_cells:
                val_str = str(c.value) if c.value is not None else ""
                # Limita tamanho para cálculo de largura
                line_len = len(val_str)
                if line_len > length:
                    length = line_len

            final_width = length + 2
            if final_width > 60:
                final_width = 60
            if final_width < 10:
                final_width = 10
            ws.column_dimensions[col_letter].width = final_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"nfse_{str(mes).zfill(2)}{ano}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
